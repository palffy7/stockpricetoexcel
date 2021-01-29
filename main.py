import datetime as dt
import pandas as pd
import pandas_datareader.data as web
from openpyxl import load_workbook

path = 'C:/Users/ralph/Documents/trading/'
filename = 'Investment.xlsx'
time = dt.datetime.combine(dt.date.today(), dt.datetime.min.time())

# read excel file from path
wb = load_workbook(path + filename)

# first sheet of file
sheet_ranges = (wb[wb.sheetnames[0]])

i = 4
first = True
while True:
    ticker = sheet_ranges['B' + str(i)].value
    if ticker is not None:
        print(ticker)
        try:
            if first:
                df = web.DataReader(ticker, 'yahoo', time)
                df['ticker'] = ticker
                first = False
            else:
                df_i = web.DataReader(ticker, 'yahoo', time)
                df_i['ticker'] = ticker
                df = df.append(df_i)
        except:
            continue
        i += 1
    else:
        break
print(df)
df.to_csv(path + 'tickerprices.csv')
# start = dt.datetime(2021, 1, 15)
# end = dt.datetime(2021, 1, 1)

#
# # df = web.DataReader('TSLA', 'yahoo', start, end)



